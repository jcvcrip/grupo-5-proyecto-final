import serial
import openpyxl
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox

# Configuración del puerto serial
ser = serial.Serial('COM5', 9600)  # Ajusta el puerto según sea necesario

# Ruta del archivo Excel
file_path = 'C:/Users/ASUS/Documents/Arduino/SENSOR_CARDIACO/tabla_de_datos_bpm_utn.xlsx'

# Crear o cargar el archivo Excel
try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Hora", "Fecha", "Nombre", "Edad", "Genero", "Codigo", "BPM Final", "Estado"])
    wb.save(file_path)

def guardar_datos(hora, fecha, nombre, edad, genero, codigo, bpm_final, estado):
    ws.append([hora, fecha, nombre.upper(), edad, genero.upper(), codigo, bpm_final, estado])
    wb.save(file_path)
    print(f"Datos guardados en: {file_path}")

def limpiar_etiquetas():
    label_codigo.config(text="Codigo: ")
    label_bpm.config(text="BPM Final: ")
    label_estado.config(text="Estado: ")

def limpiar_campos():
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_genero.delete(0, tk.END)

def procesar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    genero = entry_genero.get().upper()

    if not nombre or not edad or not genero:
        messagebox.showerror("Error", "Todos los campos son obligatorios")
        return
    
    if not edad.isdigit():
        messagebox.showerror("Error", "Edad inválida")
        return

    if genero not in ['M', 'F']:
        messagebox.showerror("Error", "Género inválido, debe ser 'M' o 'F'")
        return

    codigo_aleatorio = label_codigo.cget("text").split(": ")[1]
    bpm_final = label_bpm.cget("text").split(": ")[1]
    estado = label_estado.cget("text").split(": ")[1]

    hora_actual = datetime.now().strftime("%H:%M:%S")
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    guardar_datos(hora_actual, fecha_actual, nombre, edad, genero, codigo_aleatorio, bpm_final, estado)
    messagebox.showinfo("Datos guardados", f"Datos guardados: {hora_actual}, {fecha_actual}, {nombre}, {edad}, {genero}, {codigo_aleatorio}, {bpm_final} BPM, {estado}")
    limpiar_campos()
    limpiar_etiquetas()
    cargar_datos_excel()  # Actualizar la tabla después de guardar los datos

def actualizar_datos_sensor():
    if ser.in_waiting > 0:
        linea = ser.readline().decode('utf-8').strip()
        print(linea)
        if linea.startswith("Code:"):
            label_codigo.config(text=f"Codigo: {linea.split(': ')[1]}")
        elif linea.startswith("Final BPM:"):
            label_bpm.config(text=f"BPM Final: {linea.split(': ')[1]}")
        elif linea.startswith("Status:"):
            label_estado.config(text=f"Estado: {linea.split(': ')[1]}")
    root.after(1000, actualizar_datos_sensor)

def cargar_datos_excel():
    # Limpiar datos previos en el Treeview
    for item in tree.get_children():
        tree.delete(item)
    
    # Cargar datos desde el archivo Excel
    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)

def eliminar_fila_seleccionada():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Advertencia", "Selecciona una fila para eliminar")
        return
    
    for item in selected_item:
        item_values = tree.item(item, "values")
        if item_values:
            tree.delete(item)
            # Eliminar la fila correspondiente en el archivo Excel
            for row in ws.iter_rows(min_row=2):
                if [cell.value for cell in row] == list(item_values):
                    ws.delete_rows(row[0].row, 1)
                    wb.save(file_path)
                    break
        else:
            messagebox.showwarning("Advertencia", f"No se encontraron valores para el ítem seleccionado: {item}")

    messagebox.showinfo("Información", "Fila(s) eliminada(s) correctamente")

# Crear la interfaz gráfica
root = tk.Tk()
root.title("Ingreso de Datos")

# Tamaño de la ventana
root.geometry("800x650")

# Frame para el formulario
frame_form = tk.Frame(root, bg='white', padx=20, pady=20)  # Añadir color de fondo y relleno
frame_form.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

tk.Label(frame_form, text="Nombre:", bg='white').grid(row=0, column=0)
entry_nombre = tk.Entry(frame_form)
entry_nombre.grid(row=0, column=1)

tk.Label(frame_form, text="Edad:", bg='white').grid(row=1, column=0)
entry_edad = tk.Entry(frame_form)
entry_edad.grid(row=1, column=1)

tk.Label(frame_form, text="Genero (M/F):", bg='white').grid(row=2, column=0)
entry_genero = tk.Entry(frame_form)
entry_genero.grid(row=2, column=1)

tk.Button(frame_form, text="Guardar", command=procesar_datos).grid(row=3, columnspan=2, pady=10)

# Datos del sensor
frame_sensor = tk.Frame(root, bg='white', padx=20, pady=20)  # Añadir color de fondo y relleno
frame_sensor.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

label_codigo = tk.Label(frame_sensor, text="Codigo: ", bg='white')
label_codigo.grid(row=0, column=0)

label_bpm = tk.Label(frame_sensor, text="BPM Final: ", bg='white')
label_bpm.grid(row=1, column=0)

label_estado = tk.Label(frame_sensor, text="Estado: ", bg='white')
label_estado.grid(row=2, column=0)

# Treeview para mostrar la tabla de Excel
tree_frame = tk.Frame(root, bg='white', padx=20, pady=20)  # Añadir color de fondo y relleno
tree_frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

tree = ttk.Treeview(tree_frame, columns=("Hora", "Fecha", "Nombre", "Edad", "Genero", "Codigo", "BPM Final", "Estado"), show="headings")
tree.heading("Hora", text="Hora", anchor=tk.CENTER)
tree.heading("Fecha", text="Fecha", anchor=tk.CENTER)
tree.heading("Nombre", text="Nombre", anchor=tk.CENTER)
tree.heading("Edad", text="Edad", anchor=tk.CENTER)
tree.heading("Genero", text="Genero", anchor=tk.CENTER)
tree.heading("Codigo", text="Codigo", anchor=tk.CENTER)
tree.heading("BPM Final", text="BPM Final", anchor=tk.CENTER)
tree.heading("Estado", text="Estado", anchor=tk.CENTER)

tree.column("Hora", width=80, anchor=tk.CENTER)
tree.column("Fecha", width=100, anchor=tk.CENTER)
tree.column("Nombre", width=120, anchor=tk.CENTER)
tree.column("Edad", width=80, anchor=tk.CENTER)
tree.column("Genero", width=80, anchor=tk.CENTER)
tree.column("Codigo", width=120, anchor=tk.CENTER)
tree.column("BPM Final", width=80, anchor=tk.CENTER)
tree.column("Estado", width=120, anchor=tk.CENTER)

tree.pack(fill=tk.BOTH, expand=True)

# Botón Eliminar al final
tk.Button(root, text="Eliminar", command=eliminar_fila_seleccionada).pack(pady=10)

# Cargar los datos iniciales
cargar_datos_excel()

root.after(1000, actualizar_datos_sensor)
root.mainloop()
